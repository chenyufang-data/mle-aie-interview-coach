"""Ingest hand-collected interview experiences into rag_exp/ (plan section 7b).

Sources, all under data/interview_exp/ (gitignored - they name real people):
  - the question-frequency sheet (columns rank / question / frequency /
    interviewee / company / industry / round) - the aggregated, ranked view
  - the per-interview record (multi-question cells, one row per round)
  - optional drop folder data/interview_exp/pastes/*.txt|*.md - one
    experience post per file, for future hand-collected posts

Two stages, so the paid step is always inspected first:

  .venv\\Scripts\\python grader\\ingest_questions.py
      DRY RUN (free): parse -> normalize -> classify -> dedupe within the
      sources and against rag_ml/rag_ai, then print the report and the
      exact Claude-teacher work list with a cost estimate. Writes the full
      candidate list to data/interview_exp/ingest_preview.json for review.

  .venv\\Scripts\\python grader\\ingest_questions.py --generate --confirm
      TEACHER RUN (paid): one Claude call per eligible question generates
      the rubric (model answer, key points, common mistakes, follow-ups)
      and the English rephrasing, then appends chunks to
      rag_exp/all_chunks.jsonl. Idempotent: candidates whose id already
      exists in the bank are skipped, so an interrupted run just re-runs.

Privacy: interviewee names never leave the spreadsheets - the parser drops
the name column outright, prompts carry only question text + company /
industry / round context, and the teacher is instructed to strip any
person or current-employer reference from the rephrased question. The
built bank is itself gitignored (rag_exp/* in .gitignore): the infra plan
mounts it from private storage, never from the repository.
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR))

from coach import config  # noqa: E402

config.load_env_file()

from retrieval import tokenize  # noqa: E402

EXP_DIR = BASE_DIR / "data" / "interview_exp"
PASTE_DIR = EXP_DIR / "pastes"
PREVIEW_PATH = EXP_DIR / "ingest_preview.json"
OUT_PATH = BASE_DIR / "rag_exp" / "all_chunks.jsonl"

# The five round tags the mock simulates (plan section 7b) plus logistics,
# which is parsed and reported but never sent to the teacher: pure facts
# (visa status, location, notice period) have no gradeable answer.
MODULE_LABEL = {
    "behavioral": "HR & Behavioral",
    "experience": "Experience Deep-Dive",
    "technical": "Technical Discussion",
    "coding": "Coding",
    "system_design": "System Design",
    "logistics": "Logistics",
}

# Teacher cost estimate per question: ~700 input tokens (system + question
# + occurrence context) and ~1100 output (rubric JSON + adaptive-thinking
# overhead) at Opus-tier $5/$25 per million.
EST_IN_TOKENS, EST_OUT_TOKENS = 700, 1100
IN_PRICE, OUT_PRICE = 5e-6, 25e-6

# ---------------------------------------------------------------------------
# Parsing


def _clean(text):
    text = unicodedata.normalize("NFKC", str(text))
    return re.sub(r"\s+", " ", text).strip()


def read_frequency_sheet(path):
    """The aggregated sheet: a ranked question row followed by unranked
    occurrence rows (same question, another company). The interviewee
    column (index 3) is deliberately never read."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    candidates = []
    for row in rows:
        rank, question, freq, _name, company, industry, rnd = (
            _clean(c) if c is not None else "" for c in (list(row) + [""] * 7)[:7]
        )
        occurrence = {"company": company, "industry": industry, "round": rnd}
        if question:
            candidates.append({
                "text": question,
                "frequency": int(float(freq)) if freq else 1,
                "occurrences": [occurrence],
                "source": "frequency_sheet",
            })
        elif candidates and (company or rnd):
            candidates[-1]["occurrences"].append(occurrence)
    wb.close()
    return candidates


# Lines that are interviewer narration, not questions.
_NARRATION = re.compile(
    r"^(背景|接下来|next steps?$|公司介绍|.{0,14}介绍$|面试流程|上半部分$|下半部分$)"
)
# Section headers carry a topic for the lines that follow ("第二部分：ML
# system design"), or an inline question after the colon.
_SECTION = re.compile(r"^(第[一二三四五]部分|上半部分|下半部分)\s*[：:]\s*(.*)$")
# List prefixes like 常规面试： / 常规： / 替雇主问quick question： -
# the remainder splits on Chinese semicolons into separate items.
_LIST_PREFIX = re.compile(r"^(常规(面试)?(问题)?|.{0,12}(问|question)s?)\s*[：:]\s*(.+)$")
_ENUM = re.compile(r"^\s*(?:\d+\s*[、.．)）]+|[-•–]\s*|\(\d+\)|（\d+）)\s*")
# Spoken-transcript throat-clearing at the start of a question.
_DISCOURSE = re.compile(r"^(?:(?:and|but|so|also|okay|alright|now|um)[,，]?\s+)+",
                        re.IGNORECASE)


def split_record_cell(cell):
    """One 面试问题记录 cell -> question candidates + dropped narration.

    New questions in these cells are numbered or bulleted; an unnumbered
    line continues the previous candidate (multi-line system-design
    statements, category lists) unless a section header just reset the
    scope, in which case it starts the section's question."""
    kept, dropped = [], []
    open_new = False  # a section header makes the next unnumbered line a new candidate

    for raw in str(cell).splitlines():
        line = _clean(raw)
        if not line:
            continue

        section = _SECTION.match(line)
        if section:
            rest = _clean(section.group(2))
            open_new = True
            if rest and len(tokenize(rest)) + len(re.findall(r"[一-鿿]", rest)) >= 5:
                kept.append(rest)
                open_new = False
            continue

        listed = _LIST_PREFIX.match(line)
        if listed and ("；" in listed.group(5) or ";" in listed.group(5)):
            for part in re.split(r"[；;]", listed.group(5)):
                part = _clean(part)
                if part:
                    kept.append(part)
            open_new = False
            continue
        if listed and listed.group(1).startswith("常规"):
            kept.append(_clean(listed.group(5)))
            open_new = False
            continue

        enumerated = bool(_ENUM.match(line))
        while _ENUM.match(line):  # "-2)." stacks a bullet on a number
            line = _ENUM.sub("", line)
        line = _DISCOURSE.sub("", line)
        if not line:
            continue
        if _NARRATION.match(line) and "?" not in line and "？" not in line:
            dropped.append(line)
            continue

        if enumerated or open_new or not kept:
            kept.append(line)
            open_new = False
        else:
            kept[-1] = kept[-1] + " " + line
    return kept, dropped


def read_interview_record(path):
    """The per-interview record: one row per round; a row with an empty
    company column is a later round at the company above. Only columns
    company / title / round / question-record / aggregation-flag are read.

    Rows whose 校长整理情况 column says 已整理 were already hand-aggregated
    into the frequency sheet - that aggregation is a human dedupe (it
    condensed e.g. an 18-probe serving deep-dive into one ranked
    question), so those cells are trusted and skipped rather than
    re-parsed into near-duplicates."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    candidates, dropped = [], []
    skipped_cells = 0
    company = title = ""
    for row in rows:
        cells = (list(row) + [None] * 10)[:10]
        company = _clean(cells[1]) if cells[1] else company
        title = _clean(cells[3]) if cells[3] else title
        rnd = _clean(cells[6]) if cells[6] else ""
        # The record cell keeps its raw newlines - they are the question
        # boundaries split_record_cell works from.
        record = str(cells[8]) if cells[8] else ""
        if not record.strip():
            continue
        if cells[9] and _clean(cells[9]):
            skipped_cells += 1
            continue
        kept, gone = split_record_cell(record)
        dropped.extend(gone)
        for text in kept:
            candidates.append({
                "text": text,
                "frequency": 1,
                "occurrences": [{"company": company, "role_title": title, "round": rnd}],
                "source": "interview_record",
            })
    wb.close()
    return candidates, dropped, skipped_cells


def read_pastes():
    """Future hand-collected posts: drop a .txt/.md per experience into
    data/interview_exp/pastes/. Every non-empty line that survives the
    same narration filter becomes a candidate."""
    candidates, dropped = [], []
    if not PASTE_DIR.is_dir():
        return candidates, dropped
    for path in sorted(PASTE_DIR.glob("*.txt")) + sorted(PASTE_DIR.glob("*.md")):
        kept, gone = split_record_cell(path.read_text(encoding="utf-8"))
        dropped.extend(gone)
        for text in kept:
            candidates.append({
                "text": text,
                "frequency": 1,
                "occurrences": [{"company": "", "round": ""}],
                "source": f"paste:{path.name}",
            })
    return candidates, dropped


# ---------------------------------------------------------------------------
# Classification

_PATTERNS = [
    ("coding", r"coding|solve\s*:|implement|write (a|from)|leetcode|from scratch|白板.*(cod|题)"),
    ("system_design", r"system design|design.*\b(system|pipeline|ranking)|how would you (design|build)|架构"),
    ("logistics", r"visa|spons|relocat|on[- ]?site|hybrid|notice\b|"
                  r"located|location|where.*based|based (out|in)|base在|your base\b|"
                  r"calling in from|office|days a week|availability|start date|"
                  r"schedul|consent|authoriz|work author|move immediat|签证|入职|办公"),
    ("experience", r"walk me through|tell me about (a|the|your|more)|experience with|"
                   r"have you (work|done|built|fine-tun|tried|isolat)|production issue|"
                   r"debug|proud of|challeng|你.*(经历|经验)|有没有|有什么.*经验"),
    # Short acronyms carry \b on both sides: "iou" lives inside "curious",
    # "sla" inside "translate", "api" inside "rapidly".
    ("technical", r"precision|recall|overfit|underfit|gradient|pytorch|tensor|"
                  r"model|deploy|serv(ed|ing)|monitor|roll.*back|kubernetes|"
                  r"terraform|feature|metric|fine-?tun|token|transformer|neural|"
                  r"statis|compiled|database|latency|observability|privacy|"
                  r"\b(sft|rlhf|rag|llm|k8s|iou|r2|r²|oop|api|sla|aws|gpu|cuda|pinn)\b"),
]

# Conversational fragments that only make sense inside their interview -
# they reference the candidate's previous answer, not a standalone
# question. The teacher's followups field is where this material lives.
_FOLLOWUP = re.compile(
    r"is that (correct|right)|^(it )?sounds like\b|"
    r"\byou (said|mentioned|just (said|mentioned|trained))\b|"
    r"\b(this|that) (code|project)\b|"
    r"\bbeen there\b|\bdo (you|they) do (there|exactly)\b")


def classify(text, round_hint=""):
    """round_hint is unused on purpose: behavioral questions get asked in
    technical rounds too, so the text decides and behavioral is the
    default. The teacher assigns the final round tag anyway."""
    del round_hint
    lowered = text.lower()
    if _FOLLOWUP.search(lowered):
        return "followup"
    for category, pattern in _PATTERNS:
        if re.search(pattern, lowered):
            return category
    return "behavioral"


def language_of(text):
    cjk = len(re.findall(r"[一-鿿]", text))
    if not cjk:
        return "en"
    return "zh" if cjk > len(text) / 3 else "mixed"


def too_vague(text):
    """Terse memos ("PyTorch 需求") can't yield a faithful rubric; they are
    reported for the user to expand by hand, not sent to the teacher. A
    question mark or an imperative task verb rescues a short line
    ("implement stack from scratch" is a complete coding prompt)."""
    if "?" in text or "？" in text:
        return False
    if re.match(r"(implement|write|solve|design|build|draft|explain)\b", text.lower()):
        return False
    weight = len(tokenize(text)) + len(re.findall(r"[一-鿿]", text)) // 2
    return weight < 5


# ---------------------------------------------------------------------------
# Dedupe

_STOP = {"a", "an", "the", "you", "your", "do", "did", "have", "has", "what",
         "whats", "is", "are", "was", "were", "how", "why", "me", "my", "i",
         "we", "our", "to", "of", "in", "on", "for", "and", "or", "about",
         "tell", "can", "could", "would", "any", "with", "that", "this", "it"}


def _content_tokens(text):
    return {t for t in tokenize(text) if t not in _STOP}


def _containment(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / min(len(a), len(b))


def merge_duplicates(candidates, threshold=0.65):
    """The frequency sheet comes first (already aggregated, ranked); record
    and paste lines that lexically contain / are contained by an accepted
    question merge into it, pooling occurrences. Frequencies from the
    sheet are authoritative - merged record lines don't inflate them,
    since the sheet was aggregated from those same records."""
    accepted, merges = [], []
    for cand in candidates:
        tokens = _content_tokens(cand["text"])
        best, best_score = None, 0.0
        for acc in accepted:
            score = _containment(tokens, acc["_tokens"])
            if score > best_score:
                best, best_score = acc, score
        if best is not None and best_score >= threshold:
            best["occurrences"].extend(
                o for o in cand["occurrences"] if o not in best["occurrences"])
            if cand["source"] == best["source"] == "frequency_sheet":
                best["frequency"] += cand["frequency"]
            merges.append({"text": cand["text"], "into": best["text"],
                           "score": round(best_score, 2)})
        else:
            cand["_tokens"] = tokens
            accepted.append(cand)
    for acc in accepted:
        del acc["_tokens"]
    return accepted, merges


# HR-screen questions are a small closed set of intents asked in wildly
# different words ("Why are you choosing to leave Glodon?" / "What's
# bringing you to Niagara?" / "为什么想离开当前的公司") - exactly what
# lexical containment misses. Candidates in the same intent merge into the
# first one seen (the frequency sheet is processed first, so its ranked
# phrasing is the canon). Applied only to behavioral/experience candidates;
# technical questions never merge on intent.
_INTENTS = [
    ("compensation", r"compensation|salary|薪酬|base value|rough estimate"),
    ("why_us", r"bringing you to|motivat\w*.{0,25}apply|"
               r"interested in (a|the|this|our).{0,20}compan|why (us|this company)"),
    ("pipeline_status", r"interviewing (elsewhere|with)|other (ongoing )?interview|"
                        r"final stages|actively (looking|interviewing|searching)|"
                        r"job search|career search|(your|the) timeline|making a decision|"
                        r"how far along|passively or actively|on ?going oppo|"
                        r"talked to any other company"),
    ("tell_me_about_yourself", r"tell (me|my) .{0,20}(yourself|your background|your cv|"
                               r"your experience|resume)|overview of your (cv|career)|"
                               r"walk me through your resume|highlight your background|"
                               r"^introduce yourself|个人介绍|自我介绍"),
    ("why_leaving", r"(why|reason|what.?s bringing|motivat\w*).{0,45}(leav|switch|"
                    r"change|new (role|opportunit|job)|apply|explore|"
                    r"opportunit\w* outside)|looking to leave|"
                    r"why are you looking|searching for new|为什么.{0,10}离开"),
    ("questions_for_me", r"questions? (for (me|us|you)|unclear|at all)|"
                         r"(you )?have any (initial |other |specific )?questions|"
                         r"anything (else|i can help|i didn.?t touch)|help clarify|"
                         r"before (we wrap|i let you go)"),
    ("know_our_company", r"(know|familiar|heard) (much )?about|were you familiar|"
                         r"公司.{0,8}(熟悉|了解)"),
    ("current_company", r"(about|introduce|explain|tell me about) .{0,15}(current|your) "
                        r"(company|employer|position|role)|is \w+ a startup|"
                        r"what.?s the company doing|介绍当前雇主|"
                        r"(当前|现在)的?(公司|雇主|职责)"),
    ("next_role_wants", r"(what|things|important) .{0,40}(looking for|important (to|for) "
                        r"you).{0,40}(next|role|opportunit|company|compan)|"
                        r"kind of compan(y|ies).{0,45}looking|top things|"
                        r"differentiating between one opportunit"),
]


def intent_of(text):
    lowered = text.lower()
    for intent, pattern in _INTENTS:
        if re.search(pattern, lowered):
            return intent
    return None


def merge_by_intent(candidates):
    canon, merges = {}, []
    kept = []
    for cand in candidates:
        if cand["category"] not in ("behavioral", "experience"):
            kept.append(cand)
            continue
        intent = intent_of(cand["text"])
        if intent is None:
            kept.append(cand)
            continue
        anchor = canon.get(intent)
        if anchor is None:
            cand["intent"] = intent
            canon[intent] = cand
            kept.append(cand)
        else:
            anchor["occurrences"].extend(
                o for o in cand["occurrences"] if o not in anchor["occurrences"])
            merges.append({"text": cand["text"], "into": anchor["text"],
                           "intent": intent})
    return kept, merges


def bank_overlap(candidates, drop_threshold=0.8, note_threshold=0.6):
    """Compare against rag_ml/rag_ai. Near-exact matches drop (the course
    bank already serves them); topical overlaps only get a bank_ref note -
    "this was actually asked at company X" is the exp bank's whole value."""
    bank = []
    for path in config.CORPUS_PATHS.values():
        if path.exists():
            with path.open(encoding="utf-8") as handle:
                bank.extend(json.loads(line) for line in handle if line.strip())
    bank_tokens = [
        (chunk["id"], _content_tokens(
            chunk["interview"]["question"] + " " + " ".join(chunk["interview"]["key_points"])))
        for chunk in bank
    ]
    dropped = []
    for cand in candidates:
        tokens = _content_tokens(cand["text"])
        best_id, best_score = None, 0.0
        for chunk_id, chunk_tokens in bank_tokens:
            score = _containment(tokens, chunk_tokens)
            if score > best_score:
                best_id, best_score = chunk_id, score
        if best_score >= drop_threshold:
            cand["bank_dup"] = best_id
            dropped.append(cand)
        elif best_score >= note_threshold:
            cand["bank_ref"] = best_id
    return [c for c in candidates if "bank_dup" not in c], dropped


# ---------------------------------------------------------------------------
# Teacher (stage B)

TEACHER_SYSTEM = """You are a senior MLE/AIE interviewer writing grading \
rubrics for real interview questions gathered from actual interviews. For \
each question you receive, produce the rubric a coach needs to grade a \
candidate's spoken answer.

Rules:
- The rephrased question is in natural interviewer English, faithful to \
what was asked. If the original is Chinese or mixed, translate it.
- Generalize away anything tied to one candidate: never include any \
person's name, and replace references to a specific current employer with \
"your current company". Keep the INTERVIEWING company's domain when it \
shapes the question (e.g. a legal-LLM startup asking about document AI).
- model_answer is what a strong candidate actually says: concrete, \
first-person where natural, 3-6 sentences.
- key_points are the gradeable elements an answer must hit; \
common_mistakes are real failure modes; followups are what this \
interviewer would probe next."""

RUBRIC_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "question": {"type": "string"},
        "topic": {"type": "string"},
        "tags": {"type": "array", "items": {"type": "string"}},
        "difficulty": {"type": "string",
                       "enum": ["beginner", "beginner-intermediate",
                                "intermediate", "advanced"]},
        "round": {"type": "string",
                  "enum": ["behavioral", "experience", "technical",
                           "coding", "system_design"]},
        "model_answer": {"type": "string"},
        "key_points": {"type": "array", "items": {"type": "string"}},
        "common_mistakes": {"type": "array", "items": {"type": "string"}},
        "followups": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["question", "topic", "tags", "difficulty", "round",
                 "model_answer", "key_points", "common_mistakes", "followups"],
}


def teacher_prompt(cand):
    occurrences = [
        " / ".join(filter(None, [o.get("company"), o.get("industry"),
                                 o.get("role_title"), o.get("round")]))
        for o in cand["occurrences"]
    ]
    context = "\n".join(f"- {line}" for line in occurrences if line) or "- (unknown)"
    return (
        f"Question as recorded (verbatim):\n{cand['text']}\n\n"
        f"Asked {cand['frequency']} time(s); where (company / industry / role / round):\n"
        f"{context}\n\n"
        f"Heuristic round guess: {cand['category']} (override if wrong).\n"
        "Write the rubric."
    )


def generate_rubric(cand):
    from coach import llm

    response = llm.get_client().messages.create(
        model=os.environ.get("ANTHROPIC_MODEL", config.DEFAULT_MODEL),
        max_tokens=8000,
        system=TEACHER_SYSTEM,
        thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": teacher_prompt(cand)}],
        output_config={"format": {"type": "json_schema", "schema": RUBRIC_SCHEMA}},
    )
    if response.stop_reason == "refusal":
        raise RuntimeError("teacher declined this question")
    text = next((b.text for b in response.content if b.type == "text"), "")
    return json.loads(text)


def candidate_id(index, text):
    slug = "_".join(tokenize(text)[:4]) or "question"
    return f"exp_{index:03d}_{slug}"


def build_chunk(cand, rubric):
    companies = sorted({o.get("company") for o in cand["occurrences"] if o.get("company")})
    industries = sorted({o.get("industry", "") for o in cand["occurrences"] if o.get("industry")})
    roles = sorted({o.get("role_title", "") for o in cand["occurrences"] if o.get("role_title")})
    return {
        "id": cand["id"],
        "interview": {
            "question": rubric["question"],
            "model_answer": rubric["model_answer"],
            "key_points": rubric["key_points"],
            "common_mistakes": rubric["common_mistakes"],
            "followups": rubric["followups"],
        },
        "metadata": {
            "module": MODULE_LABEL[rubric["round"]],
            "topic": rubric["topic"],
            "tags": rubric["tags"],
            "difficulty": rubric["difficulty"],
            "round": rubric["round"],
            "frequency": cand["frequency"],
            "companies": companies,
            "industries": industries,
            "role_family": roles or ["MLE"],
            "language": language_of(cand["text"]),
            "original": cand["text"],
            "source": cand["source"],
            **({"bank_ref": cand["bank_ref"]} if "bank_ref" in cand else {}),
        },
    }


# ---------------------------------------------------------------------------
# Pipeline + report


def collect():
    freq_path = next(EXP_DIR.glob("*频率*.xls*"), None)
    record_path = next((p for p in EXP_DIR.glob("*.xlsx") if p != freq_path), None)
    if freq_path is None and record_path is None:
        raise SystemExit(f"no .xlsx sources found under {EXP_DIR}")

    candidates, dropped = [], []
    skipped_cells = 0
    if freq_path is not None:
        candidates.extend(read_frequency_sheet(freq_path))
    if record_path is not None:
        got, gone, skipped_cells = read_interview_record(record_path)
        candidates.extend(got)
        dropped.extend(gone)
    got, gone = read_pastes()
    candidates.extend(got)
    dropped.extend(gone)

    candidates, merges = merge_duplicates(candidates)
    for cand in candidates:
        cand["category"] = classify(cand["text"])
        cand["language"] = language_of(cand["text"])
        cand["vague"] = too_vague(cand["text"])
    candidates, intent_merges = merge_by_intent(candidates)
    merges.extend(intent_merges)
    candidates, bank_dups = bank_overlap(candidates)

    for index, cand in enumerate(
            (c for c in candidates
             if c["category"] not in ("logistics", "followup") and not c["vague"]),
            start=1):
        cand["id"] = candidate_id(index, cand["text"])
    return candidates, merges, bank_dups, dropped, skipped_cells


def eligible(candidates):
    return [c for c in candidates if "id" in c]


def print_report(candidates, merges, bank_dups, dropped, skipped_cells):
    work = eligible(candidates)
    by_category = Counter(c["category"] for c in candidates)
    print(f"candidates after dedupe: {len(candidates)} "
          f"(merged {len(merges)} duplicates; {len(bank_dups)} already in "
          f"rag_ml/rag_ai; {len(dropped)} narration lines dropped; "
          f"{skipped_cells} record cells already hand-aggregated into the "
          f"frequency sheet, skipped)")
    print("by category:", dict(by_category.most_common()))
    skipped_vague = [c for c in candidates
                     if c["vague"] and c["category"] not in ("logistics", "followup")]
    print(f"teacher work list: {len(work)} questions "
          f"(logistics {by_category.get('logistics', 0)}, "
          f"context-bound followups {by_category.get('followup', 0)} and "
          f"too-vague {len(skipped_vague)} excluded)")
    cost = len(work) * (EST_IN_TOKENS * IN_PRICE + EST_OUT_TOKENS * OUT_PRICE)
    print(f"estimated teacher cost: ~${cost:.2f} "
          f"({len(work)} calls x ~{EST_IN_TOKENS} in / ~{EST_OUT_TOKENS} out tokens)")
    print()
    for cand in work:
        marker = " [bank_ref]" if "bank_ref" in cand else ""
        translate = " [translate]" if cand["language"] != "en" else ""
        print(f"  {cand['id']}  ({cand['category']}, freq {cand['frequency']})"
              f"{marker}{translate}")
        print(f"      {cand['text'][:110]}")
    if skipped_vague:
        print("\ntoo vague for a faithful rubric (expand by hand in the sheet to include):")
        for cand in skipped_vague:
            print(f"  - {cand['text'][:90]}")
    followups = [c for c in candidates if c["category"] == "followup"]
    if followups:
        print("\ncontext-bound followups (live in rubrics' followups field, not as chunks):")
        for cand in followups:
            print(f"  - {cand['text'][:90]}")


def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--generate", action="store_true",
                        help="call the Claude teacher and build rag_exp/all_chunks.jsonl")
    parser.add_argument("--confirm", action="store_true",
                        help="required with --generate: acknowledge the printed cost")
    parser.add_argument("--limit", type=int, default=0,
                        help="generate only the first N eligible questions")
    args = parser.parse_args()

    candidates, merges, bank_dups, dropped, skipped_cells = collect()
    print_report(candidates, merges, bank_dups, dropped, skipped_cells)

    PREVIEW_PATH.parent.mkdir(parents=True, exist_ok=True)
    preview = {
        "candidates": [{k: v for k, v in c.items()} for c in candidates],
        "merges": merges,
        "bank_duplicates": [{"text": c["text"], "bank_dup": c["bank_dup"]} for c in bank_dups],
        "dropped_narration": dropped,
    }
    PREVIEW_PATH.write_text(
        json.dumps(preview, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\npreview written to {PREVIEW_PATH.relative_to(BASE_DIR)}")

    if not args.generate:
        print("dry run only - re-run with --generate --confirm to build the bank")
        return
    if not args.confirm:
        raise SystemExit("--generate spends real Claude tokens: re-run with --confirm "
                         "after checking the estimate above")

    existing = set()
    if OUT_PATH.exists():
        with OUT_PATH.open(encoding="utf-8") as handle:
            existing = {json.loads(line)["id"] for line in handle if line.strip()}

    work = [c for c in eligible(candidates) if c["id"] not in existing]
    if args.limit:
        work = work[:args.limit]
    if not work:
        print("nothing to generate - every candidate is already in the bank")
        return

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"\ngenerating {len(work)} rubrics ({len(existing)} already in the bank)...")
    with OUT_PATH.open("a", encoding="utf-8") as handle:
        for done, cand in enumerate(work, start=1):
            try:
                rubric = generate_rubric(cand)
            except Exception as exc:  # keep going; the re-run picks up the rest
                print(f"  {cand['id']}: FAILED ({exc})")
                continue
            handle.write(json.dumps(build_chunk(cand, rubric), ensure_ascii=False) + "\n")
            handle.flush()
            print(f"  [{done}/{len(work)}] {cand['id']} -> {rubric['round']}")
    print(f"\nbank written to {OUT_PATH.relative_to(BASE_DIR)}")


if __name__ == "__main__":
    main()
